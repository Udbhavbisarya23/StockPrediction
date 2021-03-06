import os
import openpyxl
import xlsxwriter
import yfinance as yf
from pathlib import Path


xlsx_file = os.path.join(os.path.dirname(__file__), 'Graphs/Bipartite_Graph/Bipartite.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 

# Read the active sheet:
sheet = wb_obj.active

# Shareholder array contains the names of the shareholders
shareholders = []
for column in sheet.iter_cols(2, sheet.max_column):
    shareholders.append(column[0].value)

# Stocks array contains the names of the stocks
stocks=[]
for row in sheet.iter_rows(2, sheet.max_row):
    stocks.append(row[0].value)

# Reading the bipartite graph from the excel file
bipartite_graph=[[None for i in range(len(shareholders))] for j in range(len(stocks))]

# Shareholder dictionary contains a mapping from stocks to their top 10 shareholders
shareholder={stock:[] for stock in range(len(stocks))}
for row in range(2, sheet.max_row+1):
    for column in range(2, sheet.max_column+1):
        value=sheet.cell(row,column).value
        bipartite_graph[row-2][column-2]=value
        if value!=0:
            shareholder[row-2].append(column-2)

# Obtaining the Industry of each stock 
industries = {}
for stock in stocks:
    try:
        industries[stock] = yf.Ticker(stock).info['sector']
    except:
        industries[stock] = "NA"


#Single_mode contains the new single mode projection of the graph
vari = 0
single_mode=[[0 for i in range(len(stocks))] for j in range(len(stocks))]
for current_stock in range(len(stocks)):
    for second_stock in range(current_stock+1,len(stocks)):
        value=0
        count = 0
        for holder in shareholder[current_stock]:
            if holder in shareholder[second_stock]:
                value += bipartite_graph[current_stock][holder]
                value += bipartite_graph[second_stock][holder]
                count += 1
        value=value/2

        if count > 9 and industries[stocks[current_stock]] == industries[stocks[second_stock]] and industries[stocks[current_stock]] != "NA":
            value = 1
        elif count > 9:
            value = value
        elif count <= 9:
            value = 0

        single_mode[current_stock][second_stock]=value
        single_mode[second_stock][current_stock]=value

# Creating an Excel Sheet for saving Corporation network
wb = xlsxwriter.Workbook(os.path.join(os.path.dirname(__file__), 'Graphs/Corporation_Network/Corporation_network_count9.xlsx'))
sheet = wb.add_worksheet()
for i in range(len(stocks)):
    sheet.write(0, i+1, stocks[i])
for i in range(len(stocks)):
    sheet.write(i+1, 0, stocks[i])
for i in range(len(stocks)):
    for j in range(len(stocks)):
        sheet.write(i+1, j+1, single_mode[i][j])
wb.close()

print("Corporation network ready")
    

